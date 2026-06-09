#!/usr/bin/env python3
"""Build the top-50 medical-school expansion target list from BRIMR data."""

from __future__ import annotations

import csv
import hashlib
import json
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ModuleNotFoundError as exc:
    raise SystemExit(
        "Missing openpyxl. Install dependencies with: "
        "python3 -m pip install --target /tmp/penn_corpus_deps -r requirements.txt"
    ) from exc


ROOT = Path(__file__).resolve().parents[1]
ARTIFACTS = ROOT / "artifacts" / "data"
SOURCE = ARTIFACTS / "raw" / "brimr" / "MedicalSchoolsOnly_2024_C.xlsx"
JSON_OUT = ARTIFACTS / "top50_medical_school_targets.json"
CSV_OUT = ARTIFACTS / "top50_medical_school_targets.csv"
SUMMARY_OUT = ARTIFACTS / "top50_medical_school_targets_summary.json"
GENERATED_AT = datetime.now(timezone.utc).isoformat()


def norm(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def money(value) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def sha256_json(value: object) -> str:
    encoded = json.dumps(value, ensure_ascii=True, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"Missing BRIMR workbook: {SOURCE}")

    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    ws = wb.active
    headers = [norm(cell) for cell in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]
    header_index = {header: index for index, header in enumerate(headers)}
    required = [
        "ATTRIBUTED TO MEDICAL SCHOOL",
        "MEDICAL SCHOOL LOCATION",
        "FUNDING",
        "CITY",
        "STATE OR COUNTRY NAME",
        "ORGANIZATION NAME",
        "INSTITUTION TYPE",
    ]
    missing = [header for header in required if header not in header_index]
    if missing:
        raise SystemExit(f"BRIMR workbook missing expected columns: {missing}")

    grouped: dict[str, dict] = {}
    organizations: dict[str, set[str]] = defaultdict(set)
    for row in ws.iter_rows(min_row=3, values_only=True):
        if norm(row[header_index["ATTRIBUTED TO MEDICAL SCHOOL"]]).upper() != "Y":
            continue
        school = norm(row[header_index["MEDICAL SCHOOL LOCATION"]])
        if not school:
            continue
        record = grouped.setdefault(
            school,
            {
                "school_name": school,
                "brimr_2024_total_funding": 0.0,
                "award_rows": 0,
                "city": norm(row[header_index["CITY"]]),
                "state_or_country": norm(row[header_index["STATE OR COUNTRY NAME"]]),
                "institution_type": norm(row[header_index["INSTITUTION TYPE"]]),
            },
        )
        record["brimr_2024_total_funding"] += money(row[header_index["FUNDING"]])
        record["award_rows"] += 1
        org = norm(row[header_index["ORGANIZATION NAME"]])
        if org:
            organizations[school].add(org)

    ranked = sorted(
        grouped.values(),
        key=lambda item: (-item["brimr_2024_total_funding"], item["school_name"]),
    )
    rows = []
    for rank, record in enumerate(ranked[:50], start=1):
        school = record["school_name"]
        rows.append(
            {
                "target_rank": rank,
                "school_name": school,
                "school_key": "brimr_2024_school_%03d" % rank,
                "target_basis": "BRIMR 2024 NIH funding - Schools of Medicine",
                "target_source_url": "https://brimr.org/brimr-rankings-of-nih-funding-in-2024/",
                "source_workbook": str(SOURCE.relative_to(ROOT)),
                "brimr_2024_total_funding": round(record["brimr_2024_total_funding"], 2),
                "award_rows": record["award_rows"],
                "city": record["city"],
                "state_or_country": record["state_or_country"],
                "institution_type": record["institution_type"],
                "observed_organization_names": sorted(organizations[school]),
                "gbrain_readiness_status": "not_submitted",
                "school_expansion_status": "not_started",
                "generated_at": GENERATED_AT,
            }
        )

    payload = {
        "generated_at": GENERATED_AT,
        "source": {
            "name": "BRIMR Rankings of NIH Funding in 2024 - Medical Schools Only",
            "url": "https://brimr.org/brimr-rankings-of-nih-funding-in-2024/",
            "workbook": str(SOURCE.relative_to(ROOT)),
            "methodology_url": "https://brimr.org/notes-on-methodology-2024/",
        },
        "selection_policy": (
            "Initial top-50 target set uses BRIMR 2024 total NIH funding attributed to Schools of Medicine. "
            "This is a transparent operational ranking for research-intensive medical schools; gbrain/user may "
            "approve, replace, or supplement it before declaring project-level readiness."
        ),
        "targets": rows,
    }
    JSON_OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=True, sort_keys=True) + "\n", encoding="utf-8")

    with CSV_OUT.open("w", newline="", encoding="utf-8") as handle:
        fields = [
            "target_rank",
            "school_name",
            "school_key",
            "target_basis",
            "brimr_2024_total_funding",
            "award_rows",
            "city",
            "state_or_country",
            "institution_type",
            "gbrain_readiness_status",
            "school_expansion_status",
            "target_source_url",
            "source_workbook",
            "generated_at",
        ]
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)

    rowset_material = [
        {
            "target_rank": row["target_rank"],
            "school_key": row["school_key"],
            "school_name": row["school_name"],
            "brimr_2024_total_funding": row["brimr_2024_total_funding"],
            "award_rows": row["award_rows"],
            "target_basis": row["target_basis"],
        }
        for row in rows
    ]
    summary = {
        "generated_at": GENERATED_AT,
        "targets": len(rows),
        "csv": str(CSV_OUT.relative_to(ROOT)),
        "json": str(JSON_OUT.relative_to(ROOT)),
        "source_workbook": str(SOURCE.relative_to(ROOT)),
        "source_workbook_publication_policy": "raw workbook is not committed; place it under artifacts/data/raw/brimr/ to regenerate",
        "target_basis": "BRIMR 2024 NIH funding - Schools of Medicine",
        "target_source_url": "https://brimr.org/brimr-rankings-of-nih-funding-in-2024/",
        "selection_policy": payload["selection_policy"],
        "rowset_sha256": sha256_json(rowset_material),
        "top_10_school_names": [row["school_name"] for row in rows[:10]],
        "mutation_allowed": False,
        "policy": "Top-50 target registry only. It defines the operating target set and authorizes no roster ingestion, school verification, denominator closure, URL rewrite, enrichment acceptance, or identity collapse.",
    }
    SUMMARY_OUT.write_text(json.dumps(summary, indent=2, ensure_ascii=True, sort_keys=True) + "\n", encoding="utf-8")

    print(json.dumps({"targets": len(rows), "json": str(JSON_OUT), "csv": str(CSV_OUT), "summary": str(SUMMARY_OUT)}, sort_keys=True))


if __name__ == "__main__":
    main()
